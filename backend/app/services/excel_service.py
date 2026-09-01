"""
Excel Report Service — Multi-sheet corporate Excel workbook generation using openpyxl.
"""

from io import BytesIO
from datetime import datetime, timezone
from uuid import UUID
from sqlalchemy import select, func, or_, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.task import Task
from app.models.category import Category
from app.models.audit import AuditLog
from app.models.user import User
from app.core.constants import TASK_STATUS_CONFIG, TASK_PRIORITY_CONFIG


class ExcelReportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def generate_team_report(self, team_id: UUID) -> BytesIO:
        """
        Generate a multi-sheet corporate Excel report workbook containing:
        1. Executive Summary
        2. All Team Tasks
        3. Category Statistics
        4. Audit Log History
        """
        wb = openpyxl.Workbook()

        # Define Styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Dark Slate
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="0F172A")
        subtitle_font = Font(name="Calibri", size=10, italic=True, color="64748B")
        section_font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)

        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        # -------------------------------------------------------------
        # SHEET 1: Executive Summary
        # -------------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary["A1"] = "CORPORATE WORKFLOW MANAGEMENT SYSTEM"
        ws_summary["A1"].font = title_font

        ws_summary["A2"] = f"Shared Team Workspace Executive Report — Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws_summary["A2"].font = subtitle_font

        # Query metrics
        total_q = select(func.count(Task.id)).where(Task.team_id == team_id)
        total_tasks = (await self.db.execute(total_q)).scalar_one()

        comp_q = select(func.count(Task.id)).where(Task.team_id == team_id, Task.status == "COMPLETED")
        completed_tasks = (await self.db.execute(comp_q)).scalar_one()

        pend_q = select(func.count(Task.id)).where(Task.team_id == team_id, Task.status.in_(["NEW", "IN_PROGRESS", "PENDING"]))
        pending_tasks = (await self.db.execute(pend_q)).scalar_one()

        now = datetime.now(timezone.utc)
        overdue_q = select(func.count(Task.id)).where(
            Task.team_id == team_id,
            Task.due_date.isnot(None),
            Task.due_date < now,
            Task.status != "COMPLETED",
            Task.status != "CANCELLED",
        )
        overdue_tasks = (await self.db.execute(overdue_q)).scalar_one()

        completion_rate = round((completed_tasks / total_tasks * 100.0), 1) if total_tasks > 0 else 0.0

        # Metrics Table
        ws_summary["A4"] = "KEY PERFORMANCE INDICATORS (KPIs)"
        ws_summary["A4"].font = section_font

        kpi_headers = ["Metric Parameter", "Value", "Notes"]
        for col_num, h in enumerate(kpi_headers, 1):
            cell = ws_summary.cell(row=5, column=col_num, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        kpi_rows = [
            ("Total Team Tasks", total_tasks, "Shared team workspace total"),
            ("Active Pending Tasks", pending_tasks, "NEW, IN_PROGRESS, or PENDING"),
            ("Completed Tasks", completed_tasks, "Successfully closed workflows"),
            ("Overdue Tasks", overdue_tasks, "Target due date passed"),
            ("Overall Completion Rate", f"{completion_rate}%", "Completed / Total"),
        ]

        for r_idx, row in enumerate(kpi_rows, 6):
            for c_idx, val in enumerate(row, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
                cell.font = bold_font if c_idx == 1 else regular_font
                cell.border = thin_border

        # -------------------------------------------------------------
        # SHEET 2: All Team Tasks
        # -------------------------------------------------------------
        ws_tasks = wb.create_sheet(title="All Team Tasks")
        ws_tasks.views.sheetView[0].showGridLines = True

        task_headers = [
            "Task Number",
            "Task Title",
            "Category Code",
            "Category Name",
            "Priority",
            "Status",
            "Progress %",
            "Creator User",
            "Assigned User",
            "Created Date",
            "Due Date",
            "Completed Date",
        ]

        for col_num, h in enumerate(task_headers, 1):
            cell = ws_tasks.cell(row=1, column=col_num, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if "Date" in h or "%" in h else "left", vertical="center")

        tasks_q = (
            select(Task)
            .where(Task.team_id == team_id)
            .options(
                selectinload(Task.category),
                selectinload(Task.creator),
                selectinload(Task.assignee),
            )
            .order_by(Task.created_at.desc())
        )
        tasks = (await self.db.execute(tasks_q)).scalars().all()

        for r_idx, t in enumerate(tasks, 2):
            created_str = t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else ""
            due_str = t.due_date.strftime("%Y-%m-%d") if t.due_date else ""
            completed_str = t.completed_at.strftime("%Y-%m-%d %H:%M") if t.completed_at else ""

            row_data = [
                t.task_number,
                t.title,
                t.category.category_code if t.category else "",
                t.category.category_name if t.category else "",
                t.priority,
                t.status,
                f"{round(t.progress_percentage)}%",
                t.creator.full_name if t.creator else "",
                t.assignee.full_name if t.assignee else "",
                created_str,
                due_str,
                completed_str,
            ]

            for c_idx, val in enumerate(row_data, 1):
                cell = ws_tasks.cell(row=r_idx, column=c_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border

        # -------------------------------------------------------------
        # SHEET 3: Category Statistics
        # -------------------------------------------------------------
        ws_cat = wb.create_sheet(title="Category Statistics")
        ws_cat.views.sheetView[0].showGridLines = True

        cat_headers = ["Category Code", "Category Name", "Total Tasks", "Completed Tasks", "Pending Tasks", "Completion Rate %"]
        for col_num, h in enumerate(cat_headers, 1):
            cell = ws_cat.cell(row=1, column=col_num, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        cat_q = select(Category).where(Category.team_id == team_id, Category.is_active == True)
        categories = (await self.db.execute(cat_q)).scalars().all()

        for r_idx, c in enumerate(categories, 2):
            tot_q = select(func.count(Task.id)).where(Task.team_id == team_id, Task.category_id == c.id)
            c_tot = (await self.db.execute(tot_q)).scalar_one()

            comp_q = select(func.count(Task.id)).where(Task.team_id == team_id, Task.category_id == c.id, Task.status == "COMPLETED")
            c_comp = (await self.db.execute(comp_q)).scalar_one()

            c_pend = c_tot - c_comp
            c_rate = round((c_comp / c_tot * 100.0), 1) if c_tot > 0 else 0.0

            row_data = [c.category_code, c.category_name, c_tot, c_comp, c_pend, f"{c_rate}%"]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_cat.cell(row=r_idx, column=c_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border

        # -------------------------------------------------------------
        # SHEET 4: Audit Log History
        # -------------------------------------------------------------
        ws_audit = wb.create_sheet(title="Audit Log History")
        ws_audit.views.sheetView[0].showGridLines = True

        audit_headers = ["Timestamp (UTC)", "User Name", "Action Type", "Entity Type", "Entity ID", "IP Address"]
        for col_num, h in enumerate(audit_headers, 1):
            cell = ws_audit.cell(row=1, column=col_num, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        audit_q = (
            select(AuditLog)
            .where(AuditLog.team_id == team_id)
            .options(selectinload(AuditLog.user))
            .order_by(AuditLog.timestamp.desc())
            .limit(200)
        )
        audit_logs = (await self.db.execute(audit_q)).scalars().all()

        for r_idx, a in enumerate(audit_logs, 2):
            ts_str = a.timestamp.strftime("%Y-%m-%d %H:%M:%S") if a.timestamp else ""
            row_data = [
                ts_str,
                a.user.full_name if a.user else "System",
                a.action_type,
                a.entity_type or "SYSTEM",
                str(a.entity_id) if a.entity_id else "",
                a.ip_address or "",
            ]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws_audit.cell(row=r_idx, column=c_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border

        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
